# -*- coding: utf-8 -*-
"""
V5.4 工程化自检脚本（小蜂学掌定制版）
共 12 项自检（原 10 项 + 新增 S11/S12 品牌注入断言）。

运行：python scripts/self_check.py
全绿即视为发布就绪。
"""
from __future__ import annotations
import sys, os, json, re, ast, platform, subprocess, tempfile
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

HERE = Path(__file__).resolve().parent.parent

# ── 品牌常量 ──
BRAND_NAME = "小蜂学掌"
BRAND_URL  = "www.xfxz123.com"
NAVY       = "1F4E79"

results: list[tuple[int, str, bool, str]] = []


def print_brand_header():
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║  🐝 小蜂学掌 × 精细化开发系统 Self Check  V5.4              ║")
    print("║  www.xfxz123.com  |  global-sales-elite-v5                  ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  正在执行 12 项工程自检，请稍候…                             ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()


def print_delivery_box(report_path: Path, n_pass: int, n_fail: int, total: int):
    grade = "A" if n_pass == total else ("B" if n_pass >= 10 else "C")
    abs_path = str(report_path.resolve())
    short = abs_path if len(abs_path) <= 52 else "..." + abs_path[-49:]
    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    if n_fail == 0:
        print("║  ✅ 小蜂学掌 Self Check — 全绿通过！                         ║")
    else:
        print("║  ⚠️  小蜂学掌 Self Check — 存在不通过项！                    ║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print(f"║  📊 评分：{grade} 级  ({n_pass}/{total} PASS){'':>36}║")
    print("║  📄 日志/报告路径：                                           ║")
    print(f"║    {short:<58}║")
    print("╠══════════════════════════════════════════════════════════════╣")
    print("║  💡 在 Windows 资源管理器地址栏粘贴路径可直接打开             ║")
    print("╚══════════════════════════════════════════════════════════════╝")


def read(p: Path) -> str:
    return p.read_text(encoding='utf-8', errors='replace')


def check(idx: int, name: str, ok: bool, detail: str = ''):
    results.append((idx, name, ok, detail))
    status = '✅' if ok else '❌'
    suffix = f'  ({detail})' if detail else ''
    print(f"  {status} [{idx:>2}] {name}{suffix}")


def run_py(*args) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, *args],
        capture_output=True, text=True, timeout=30,
        encoding='utf-8', errors='replace',
    )


def main():
    print_brand_header()
    print("=" * 62)
    print(f"  [Self Check] global-sales-elite-v5 V5.4 | 12 项断言")
    print("=" * 62)

    # ── 1. SKILL.md frontmatter ──
    try:
        skill_text = read(HERE / "SKILL.md")
        ok = ('name:' in skill_text) and ('display_name:' in skill_text)
        check(1, "SKILL.md frontmatter 合法（含 name + display_name）", ok)
    except Exception as e:
        check(1, "SKILL.md frontmatter 合法（含 name + display_name）", False, str(e))

    # ── 2. manifest.json semver ──
    try:
        m = json.loads(read(HERE / "manifest.json"))
        ver = m.get("version", "")
        ok = bool(re.match(r'^\d+\.\d+\.\d+$', ver))
        check(2, "manifest.json semver 格式", ok, f"version={ver}")
    except Exception as e:
        check(2, "manifest.json semver 格式", False, str(e))

    # ── 3. main.py --help 可运行 ──
    try:
        r = run_py(str(HERE / "main.py"), "--help")
        ok = r.returncode == 0
        check(3, "main.py 可独立运行（--help）", ok, r.stderr[:80] if not ok else '')
    except Exception as e:
        check(3, "main.py 可独立运行（--help）", False, str(e))

    # ── 4. payload.schema.json Draft7Validator ──
    try:
        from jsonschema import Draft7Validator
        schema = json.loads(read(HERE / "payload.schema.json"))
        Draft7Validator.check_schema(schema)
        check(4, "payload.schema.json 通过 Draft7Validator", True)
    except ImportError:
        check(4, "payload.schema.json 通过 Draft7Validator", False, "jsonschema 未安装")
    except Exception as e:
        check(4, "payload.schema.json 通过 Draft7Validator", False, str(e))

    # ── 5. 全部 .py 文件 UTF-8 无 BOM ──
    try:
        py_files = list(HERE.rglob("*.py"))
        bom_files = []
        for pf in py_files:
            raw = pf.read_bytes()
            if raw.startswith(b'\xef\xbb\xbf'):
                bom_files.append(pf.name)
        ok = not bom_files
        check(5, "全部 .py 文件 UTF-8 无 BOM", ok,
              f"BOM文件: {bom_files}" if bom_files else f"{len(py_files)} 个文件全部通过")
    except Exception as e:
        check(5, "全部 .py 文件 UTF-8 无 BOM", False, str(e))

    # ── 6. 路径使用 pathlib.Path ──
    try:
        main_text = read(HERE / "main.py")
        ok = 'pathlib' in main_text and 'Path' in main_text
        check(6, "路径使用 pathlib.Path", ok)
    except Exception as e:
        check(6, "路径使用 pathlib.Path", False, str(e))

    # ── 7. 异常覆盖三类退出码 ──
    try:
        main_text = read(HERE / "main.py")
        ok = all(code in main_text for code in ['sys.exit(1)', 'sys.exit(2)', 'sys.exit(3)'])
        check(7, "异常覆盖 3 类退出码（1/2/3）", ok)
    except Exception as e:
        check(7, "异常覆盖 3 类退出码（1/2/3）", False, str(e))

    # ── 8. logs/<run_id>.log 实际写入 ──
    try:
        sample = make_sample_payload()
        with tempfile.TemporaryDirectory() as td:
            payload_file = Path(td) / "sample.json"
            payload_file.write_text(json.dumps(sample, ensure_ascii=False), encoding='utf-8')
            r = run_py(str(HERE / "main.py"), str(payload_file), "--out-dir", td, "--no-word")
            logs = list(Path(td).rglob("*.log"))
            ok = bool(logs) and logs[0].stat().st_size > 0
            check(8, "logs/<run_id>.log 实际写入且非空", ok,
                  f"returncode={r.returncode}, logs={[l.name for l in logs]}")
    except Exception as e:
        check(8, "logs/<run_id>.log 实际写入且非空", False, str(e))

    # ── 9. Excel 品牌表头 + 冻结 + 列宽（V5.4 新规范）──
    try:
        from openpyxl import load_workbook
        sample = make_sample_payload()
        with tempfile.TemporaryDirectory() as td:
            payload_file = Path(td) / "sample.json"
            payload_file.write_text(json.dumps(sample, ensure_ascii=False), encoding='utf-8')
            run_py(str(HERE / "main.py"), str(payload_file), "--out-dir", td, "--no-word")
            xl_files = list(Path(td).rglob("*.xlsx"))
            assert xl_files, "没有生成 Excel"
            wb = load_workbook(xl_files[0])
            ws = wb.active
            # 行1 品牌表头（合并单元格，含品牌名）
            a1_val = str(ws['A1'].value or '')
            brand_ok = BRAND_NAME in a1_val and BRAND_URL in a1_val
            # 行3 数据表头绿色（#00B050）
            a3_color = str(ws.cell(row=3, column=1).fill.start_color.rgb or '')
            color_ok = a3_color.upper().endswith('00B050')
            # 冻结到 A4
            freeze_ok = ws.freeze_panes == 'A4'
            col_ok = bool(ws.column_dimensions['A'].width and ws.column_dimensions['X'].width)
            ok = brand_ok and color_ok and freeze_ok and col_ok
            check(9, f"Excel 品牌表头({BRAND_NAME})/绿表头/冻结A4/列宽",
                  ok, f"brand={brand_ok} color={color_ok} freeze={freeze_ok} col={col_ok}")
    except Exception as e:
        check(9, "Excel 品牌表头/绿表头/冻结A4/列宽", False, str(e))

    # ── 10. 跨平台兼容 ──
    try:
        m = json.loads(read(HERE / "manifest.json"))
        platforms = set(m.get("platform", []))
        defaults = m.get("default_out_dir", {})
        ok = platforms == {"windows", "macos", "linux"} and all(
            k in defaults for k in ["windows", "macos", "linux"]
        )
        sysname = platform.system().lower()
        check(10, "跨平台打包（三平台默认目录齐全）", ok, f"当前平台={sysname}")
    except Exception as e:
        check(10, "跨平台打包", False, str(e))

    # ── 11. Word 页眉页脚品牌注入（V5.4 新增）──
    try:
        word_text = read(HERE / "scripts" / "word_gen.py")
        checks_11 = [
            (BRAND_NAME in word_text,     f"含 {BRAND_NAME}"),
            (BRAND_URL in word_text,       f"含 {BRAND_URL}"),
            ('section.header' in word_text or '_add_header_footer' in word_text,  "含页眉代码"),
            ('section.footer' in word_text or '_add_header_footer' in word_text,  "含页脚代码"),
            ('add_brand_cover' in word_text,  "含品牌封面函数"),
            ('小蜂学掌_' in word_text,         "文件名含品牌前缀"),
        ]
        fail_11 = [label for cond, label in checks_11 if not cond]
        ok = not fail_11
        check(11, "Word 封面/页眉/页脚/文件名品牌注入", ok,
              f"缺: {fail_11}" if fail_11 else "全部到位")
    except Exception as e:
        check(11, "Word 封面/页眉/页脚/文件名品牌注入", False, str(e))

    # ── 12. 品牌通知框 + 绝对路径（V5.4 新增）──
    try:
        main_text = read(HERE / "main.py")
        checks_12 = [
            ('print_brand_startup' in main_text,   "启动通知框函数"),
            ('print_delivery_box'  in main_text,   "交付路径通知框函数"),
            ('╔══' in main_text,                   "╔══╗ 格式通知框"),
            (BRAND_NAME in main_text,              f"含 {BRAND_NAME}"),
            (BRAND_URL  in main_text,              f"含 {BRAND_URL}"),
        ]
        fail_12 = [label for cond, label in checks_12 if not cond]
        ok = not fail_12
        check(12, "启动通知框/交付路径框/品牌信息注入", ok,
              f"缺: {fail_12}" if fail_12 else "全部到位")
    except Exception as e:
        check(12, "启动通知框/交付路径框/品牌信息注入", False, str(e))

    # ── 汇总 ──
    fail = [r for r in results if not r[2]]
    n_pass = len(results) - len(fail)
    total = len(results)
    print()
    print("=" * 62)

    # 写简易报告
    report_path = HERE / "logs"
    report_path.mkdir(exist_ok=True)
    rpt = report_path / "self_check_report.json"
    rpt.write_text(json.dumps({
        "version": "5.4.0",
        "brand": BRAND_NAME,
        "pass": n_pass,
        "fail": len(fail),
        "total": total,
        "details": [{"idx": r[0], "name": r[1], "ok": r[2], "detail": r[3]} for r in results],
    }, ensure_ascii=False, indent=2), encoding='utf-8')

    print_delivery_box(rpt, n_pass, len(fail), total)

    if fail:
        for idx, name, _, detail in fail:
            print(f"  ❌ [{idx:>2}] {name}: {detail}")
        sys.exit(1)
    else:
        print(f"  自检结果: 全绿 {total}/{total} (ALL PASS)")
        sys.exit(0)


def make_sample_payload() -> dict:
    return {
        "channel": "google",
        "country": "美国",
        "segment": "全屋定制",
        "v1_assessment": {
            "name_simple": "TestCorp",
            "type": "美国测试型客户",
            "basis": "用于自检的样例数据",
            "rating": "A"
        },
        "v2_intelligence": {
            "company_full_name": "Test Corporation Inc.",
            "country": "USA",
            "website": "https://test.example.com",
            "founded": "2020",
            "hq": "San Francisco, CA",
            "social": "测试用社交资产"
        },
        "v3_insight": {
            "factory_scale": "中型",
            "market_position": "中端",
            "social_activity": "活跃",
            "sourcing_logic": "性价比",
            "tech_preference": "环保认证"
        },
        "v4_contacts": [
            {"name": "Alice", "title": "Buyer", "email": "alice@test.example.com",
             "linkedin": "https://linkedin.com/in/alice", "whatsapp": "+1-000-000-0000",
             "touchpoint": "邮件", "strategy": "提供样品"},
            {"name": "Bob", "title": "Manager", "email": "bob@test.example.com",
             "linkedin": "https://linkedin.com/in/bob", "whatsapp": "+1-000-000-0001",
             "touchpoint": "领英", "strategy": "InMail"},
            {"name": "Carol", "title": "VP", "email": "carol@test.example.com",
             "linkedin": "https://linkedin.com/in/carol", "whatsapp": "+1-000-000-0002",
             "touchpoint": "电话", "strategy": "高管直触"}
        ],
        "v5_negotiation": {"situation": "测试", "pain_point": "测试", "solution": "测试方案"},
        "v6_templates": {
            "chinese": "您好，这是用于自检的测试中文邮件，长度足够通过 schema。",
            "english": "Hello, this is a self-check test English email of sufficient length."
        },
        "v7_product_analysis": {"pricing": "中端", "preference": "本地化"},
        "v8_suggestions": ["建议1", "建议2", "建议3"]
    }


if __name__ == "__main__":
    main()
