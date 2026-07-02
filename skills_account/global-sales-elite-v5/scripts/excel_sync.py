# -*- coding: utf-8 -*-
"""
V5.4 Excel Sync（小蜂学掌定制版）
群发用关键信息总表（24 列，新增 渠道）：
渠道 | 公司英文名称 | 英文简称 | 官网 | 画像 | 国家 | 等级
| 联系人1 (姓名/职位/邮箱/LinkedIn/Tel)
| 联系人2 (...) | 联系人3 (...)
| 中文邮件模板 | 英文邮件模板

品牌注入（v5.4）：
  - 文件名前缀：小蜂学掌_
  - 行1：品牌表头（深海蓝 #1F4E79，含小蜂学掌 | www.xfxz123.com）
  - 行2：金色分隔条（#F2C94C）
  - 行3：原数据表头（绿色 #00B050，白字加粗）
  - 末行：版权表尾（浅灰 #F2F2F2）
"""
from __future__ import annotations
import sys, json, os
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 品牌常量 ──
BRAND_NAME    = "小蜂学掌"
BRAND_URL     = "www.xfxz123.com"
BRAND_HEADER  = f"{BRAND_NAME}  |  精细化开发系统 V5.4  |  {BRAND_URL}"
BRAND_COPY    = f"© {BRAND_NAME}  {BRAND_URL}  版权所有，未经授权禁止转载"
NAVY          = "1F4E79"
BRAND_GOLD    = "F2C94C"
FOOTER_GRAY   = "F2F2F2"
DATA_START    = 3   # 数据实际从第3行开始（行1=品牌表头, 行2=金色条, 行3=数据表头, 行4+=数据）

HEADERS = [
    '渠道',
    '公司英文名称', '英文简称', '官网', '画像（客户类型）', '国家', '等级',
    '联系人1-姓名', '联系人1-职位', '联系人1-邮箱', '联系人1-LinkedIn', '联系人1-WhatsApp/Tel',
    '联系人2-姓名', '联系人2-职位', '联系人2-邮箱', '联系人2-LinkedIn', '联系人2-WhatsApp/Tel',
    '联系人3-姓名', '联系人3-职位', '联系人3-邮箱', '联系人3-LinkedIn', '联系人3-WhatsApp/Tel',
    '中文邮件模板', '英文邮件模板',
]
COL_COUNT = len(HEADERS)

CHANNEL_LABEL = {
    "google":    "Google",
    "instagram": "Instagram",
    "facebook":  "Facebook",
    "linkedin":  "LinkedIn",
    "tiktok":    "TikTok",
    "manual":    "手工",
}

THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL   = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
BRAND_FILL    = PatternFill(start_color=NAVY,     end_color=NAVY,     fill_type='solid')
GOLD_FILL     = PatternFill(start_color=BRAND_GOLD, end_color=BRAND_GOLD, fill_type='solid')
FOOTER_FILL   = PatternFill(start_color=FOOTER_GRAY, end_color=FOOTER_GRAY, fill_type='solid')

COL_WIDTHS = {
    'A': 12,
    'B': 26, 'C': 16, 'D': 32, 'E': 26, 'F': 12, 'G': 10,
    'H': 16, 'I': 22, 'J': 30, 'K': 42, 'L': 20,
    'M': 16, 'N': 22, 'O': 30, 'P': 42, 'Q': 20,
    'R': 16, 'S': 22, 'T': 30, 'U': 42, 'V': 20,
    'W': 60, 'X': 60,
}

LAST_COL_LETTER = get_column_letter(COL_COUNT)   # 'X'


def _write_brand_rows(ws):
    """写品牌表头行（行1）和金色分隔条（行2）。"""
    # ── 行1：品牌表头（深海蓝）──
    ws.merge_cells(f'A1:{LAST_COL_LETTER}1')
    cell_brand = ws['A1']
    cell_brand.value = BRAND_HEADER
    cell_brand.font  = Font(bold=True, color='FFFFFF', size=12, name='Microsoft YaHei')
    cell_brand.fill  = BRAND_FILL
    cell_brand.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── 行2：金色分隔条 ──
    ws.merge_cells(f'A2:{LAST_COL_LETTER}2')
    ws['A2'].fill = GOLD_FILL
    ws.row_dimensions[2].height = 4


def _write_data_header(ws):
    """写数据表头（行3）。"""
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=DATA_START, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[DATA_START].height = 32
    ws.freeze_panes = f'A{DATA_START + 1}'   # 冻结到数据行上方


def _write_footer(ws):
    """写品牌版权表尾行（末行）。"""
    footer_row = ws.max_row + 1
    ws.merge_cells(f'A{footer_row}:{LAST_COL_LETTER}{footer_row}')
    fc = ws[f'A{footer_row}']
    fc.value = BRAND_COPY
    fc.font  = Font(color='595959', size=8.5, name='Microsoft YaHei')
    fc.fill  = FOOTER_FILL
    fc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[footer_row].height = 16


def init_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'V5客户群发总表'

    _write_brand_rows(ws)
    _write_data_header(ws)

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    return wb, ws


def _has_brand_header(ws) -> bool:
    """检测工作表是否已有品牌表头（判断 A1 是否含品牌名）。"""
    try:
        a1 = str(ws['A1'].value or '')
        return BRAND_NAME in a1
    except Exception:
        return False


def _remove_footer_if_exists(ws):
    """删除最末行（如果是表尾行），以便追加数据后重新写表尾。"""
    last_row = ws.max_row
    if last_row <= DATA_START:
        return
    last_val = str(ws.cell(row=last_row, column=1).value or '')
    if BRAND_NAME in last_val and '版权所有' in last_val:
        ws.delete_rows(last_row)


def normalize_contact(c: dict) -> dict:
    return {
        'name':     c.get('name', ''),
        'title':    c.get('title', ''),
        'email':    c.get('email', ''),
        'linkedin': c.get('linkedin', ''),
        'whatsapp': c.get('whatsapp', ''),
    }


def sync_excel(data_file: str, excel_path: str = None) -> str:
    with open(data_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # ── 文件名前缀：小蜂学掌_ ──
    excel_path = excel_path or os.environ.get('V5_EXCEL_PATH', f'./{BRAND_NAME}_V5客户群发总表.xlsx')
    ep = Path(excel_path).expanduser().resolve()
    # 如文件名不含品牌前缀，自动加上
    if not ep.name.startswith(f'{BRAND_NAME}_'):
        ep = ep.parent / f'{BRAND_NAME}_{ep.name}'
    ep.parent.mkdir(parents=True, exist_ok=True)
    excel_path = str(ep)

    if os.path.exists(excel_path):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            if not _has_brand_header(ws):
                # 旧版本无品牌表头，重建
                wb, ws = init_workbook()
            else:
                # 删除表尾，追加完后重新加
                _remove_footer_if_exists(ws)
        except Exception:
            wb, ws = init_workbook()
    else:
        wb, ws = init_workbook()

    v1 = data['v1_assessment']
    v2 = data['v2_intelligence']
    v4 = data.get('v4_contacts', [])
    v6 = data.get('v6_templates', {})

    contacts = [normalize_contact(c) for c in v4[:3]]
    while len(contacts) < 3:
        contacts.append(normalize_contact({}))

    row = [
        CHANNEL_LABEL.get(data.get('channel', 'manual'), '手工'),
        v2.get('company_full_name', ''),
        v1.get('name_simple', ''),
        v2.get('website', ''),
        v1.get('type', ''),
        v2.get('country', ''),
        v1.get('rating', ''),
    ]
    for c in contacts:
        row.extend([c['name'], c['title'], c['email'], c['linkedin'], c['whatsapp']])
    row.append(v6.get('chinese', ''))
    row.append(v6.get('english', ''))

    ws.append(row)
    new_row_idx = ws.max_row
    for cell in ws[new_row_idx]:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = BORDER

    # 邮件列高
    ws.row_dimensions[new_row_idx].height = 80

    # ── 重新写表尾 ──
    _write_footer(ws)

    wb.save(excel_path)
    abs_path = str(Path(excel_path).resolve())
    print(f"EXCEL_READY: {abs_path}")
    return abs_path


if __name__ == "__main__":
    if len(sys.argv) > 1:
        excel_path = sys.argv[2] if len(sys.argv) > 2 else None
        sync_excel(sys.argv[1], excel_path)
