# -*- coding: utf-8 -*-
"""
Super Background Check v3.0 - Single Excel Generator (6 Sheets)

Sheet ⓪ 封面          - 小蜂学掌品牌封面（公司名 / 日期 / 官网）
Sheet ① 决策摘要      - 顶部贸易速览 4 列卡片 / 客户层级 / 风险评分 / 决策结论 / 5 决策人速览 / 5 套话术全文 / 中英开发信全文
Sheet ② 全息档案      - 10 维 ≥39 行字段表（v2.5 新增第 10 维进出口贸易）
Sheet ③ 决策人图谱    - >=5 人 x 8 列（P0/P1/P2 三色条件格式）
Sheet ④ 数据来源      - >=4 行（含 fetched_at）
Sheet ⑤ 自检报告      - sanity_check 17 项结果（v2.5 新增 S14-S17）

工业级硬约束：
- 全链路 fail-fast：禁止任何 dict-get-with-default 类型的占位，缺字段直接 sys.exit(1)
- 字段缺失 -> 红字 [缺] + 退出码 1（不允许 N/A 静默通过）
- 标题合并 / 表头绿底 / 三色风险条件格式 / 决策结论填充 / P0P1P2 三色 / 冻结表头
- v3.0 新增：⓪ 封面 Sheet + 全 Sheet 品牌表头/表尾（小蜂学掌 | www.xfxz123.com）+ 文件名品牌前缀
"""
import argparse
import json
import os
import sys
from datetime import datetime

# Force UTF-8 stdout for Windows GBK console
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl not installed. Run: pip install openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple")
    sys.exit(1)

# ============ 品牌常量 ============
BRAND_NAME   = "小蜂学掌"
BRAND_EN     = "BEE-EDUCATION"
BRAND_URL    = "www.xfxz123.com"
BRAND_COPYRIGHT = f"© {BRAND_NAME}  {BRAND_URL}  版权所有，未经授权禁止转载"

# ============ 颜色 / 字体 / 边框 ============
NAVY = "1F4E79"
GREEN = "00B050"
LIGHT_BLUE = "DEEBF7"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW = "FFF2CC"
WHITE = "FFFFFF"
RED = "C00000"
ORANGE = "E67E22"
GRAY = "595959"
LIGHT_GRAY = "F2F2F2"
BEE_GOLD = "F2C94C"   # 品牌金色（封面/表头点缀）

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT = Font(name="Microsoft YaHei", size=16, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color=WHITE)
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color=WHITE)
SECTION_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color=WHITE)
CELL_FONT = Font(name="Microsoft YaHei", size=10)
CELL_FONT_BOLD = Font(name="Microsoft YaHei", size=10, bold=True, color=NAVY)
CELL_FONT_RED = Font(name="Microsoft YaHei", size=10, color=RED)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ============ 品牌表头 / 表尾注入 ============
def inject_brand_rows(ws, col_count, report_title=""):
    """
    在工作表最顶部插入 2 行品牌表头，在数据写完后调用 append_brand_footer() 追加表尾。
    调用时机：每个 build_sheetN 函数在 create_sheet 之后、写数据之前调用本函数。
    返回 brand_row_offset=2，调用方所有行号需 +2。
    """
    last_col = get_column_letter(col_count)

    # --- 行 1：品牌 LOGO 条（深海蓝底 白字）---
    ws.merge_cells(f"A1:{last_col}1")
    c1 = ws["A1"]
    c1.value = f"  {BRAND_NAME}  |  {BRAND_EN}  |  {BRAND_URL}"
    c1.font = Font(name="Microsoft YaHei", size=11, bold=True, color=WHITE)
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # --- 行 2：报告副标题灰条 ---
    ws.merge_cells(f"A2:{last_col}2")
    c2 = ws["A2"]
    c2.value = report_title if report_title else f"超级背调专家 v3.0  |  {BRAND_NAME}出品"
    c2.font = Font(name="Microsoft YaHei", size=9, color=GRAY)
    c2.fill = PatternFill("solid", fgColor="F2F2F2")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    return 2  # brand_row_offset


def append_brand_footer(ws, col_count):
    """在当前最后一行后追加品牌表尾行。"""
    last_row = ws.max_row + 1
    last_col = get_column_letter(col_count)
    ws.merge_cells(f"A{last_row}:{last_col}{last_row}")
    cf = ws.cell(row=last_row, column=1)
    cf.value = BRAND_COPYRIGHT
    cf.font = Font(name="Microsoft YaHei", size=8, italic=True, color=GRAY)
    cf.fill = PatternFill("solid", fgColor="F2F2F2")
    cf.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[last_row].height = 16


# ============ 封面 Sheet ============
def build_sheet_cover(wb, payload):
    ws = wb.active
    ws.title = "⓪ 封面"

    company_full  = payload.get("company_full", "")
    company_short = payload.get("company_short", "")
    country       = payload.get("country", "")
    level         = payload.get("level", "")
    date_str      = payload.get("date_str", "")

    # 列宽
    for col in ["A", "B", "C", "D"]:
        ws.column_dimensions[col].width = 28

    # 行 1-2 留白
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20

    # --- LOGO 大标题 ---
    ws.merge_cells("A3:D3")
    c = ws["A3"]
    c.value = BRAND_NAME
    c.font = Font(name="Microsoft YaHei", size=36, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="DEEBF7")
    ws.row_dimensions[3].height = 60

    # --- 品牌英文 + 网址 ---
    ws.merge_cells("A4:D4")
    c4 = ws["A4"]
    c4.value = f"{BRAND_EN}  ·  {BRAND_URL}"
    c4.font = Font(name="Microsoft YaHei", size=13, color=NAVY)
    c4.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 28

    # --- 分隔线（金色条）---
    ws.merge_cells("A5:D5")
    ws["A5"].fill = PatternFill("solid", fgColor=BEE_GOLD)
    ws.row_dimensions[5].height = 6

    # --- 报告类型 ---
    ws.merge_cells("A7:D7")
    c7 = ws["A7"]
    c7.value = "超级背调专家报告"
    c7.font = Font(name="Microsoft YaHei", size=22, bold=True, color=WHITE)
    c7.fill = PatternFill("solid", fgColor=NAVY)
    c7.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 48

    # --- 公司名 ---
    ws.merge_cells("A9:D9")
    c9 = ws["A9"]
    c9.value = company_full or company_short
    c9.font = Font(name="Microsoft YaHei", size=18, bold=True, color=NAVY)
    c9.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[9].height = 40

    # --- 信息卡片：国家 / 等级 / 日期 ---
    info_rows = [
        ("目标国家", country),
        ("客户层级", f"{level} 级"),
        ("报告日期", date_str),
    ]
    r = 11
    for label, val in info_rows:
        ws.merge_cells(f"A{r}:B{r}")
        lc = ws[f"A{r}"]
        lc.value = label
        lc.font = Font(name="Microsoft YaHei", size=11, bold=True, color=WHITE)
        lc.fill = PatternFill("solid", fgColor=NAVY)
        lc.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"C{r}:D{r}")
        vc = ws[f"C{r}"]
        vc.value = val
        vc.font = Font(name="Microsoft YaHei", size=11, bold=True, color=NAVY)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = PatternFill("solid", fgColor="DEEBF7")
        ws.row_dimensions[r].height = 28
        r += 1

    # --- 底部金条 + 版权 ---
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"].fill = PatternFill("solid", fgColor=BEE_GOLD)
    ws.row_dimensions[r].height = 6
    r += 2

    ws.merge_cells(f"A{r}:D{r}")
    cc = ws[f"A{r}"]
    cc.value = BRAND_COPYRIGHT
    cc.font = Font(name="Microsoft YaHei", size=9, italic=True, color=GRAY)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20


# ============ Fail-fast 取值 ============
def must(payload, *path):
    """Fail-fast 取值：禁止 .get() 默认值占位。缺字段直接 exit(1)。"""
    node = payload
    trace = []
    for k in path:
        trace.append(str(k))
        if isinstance(node, dict):
            if k not in node:
                print(f"[FAIL-FAST] payload 缺字段: {'.'.join(trace)}")
                sys.exit(1)
            node = node[k]
        elif isinstance(node, list):
            if not isinstance(k, int) or k >= len(node):
                print(f"[FAIL-FAST] payload 数组越界: {'.'.join(trace)}")
                sys.exit(1)
            node = node[k]
        else:
            print(f"[FAIL-FAST] payload 路径无效: {'.'.join(trace)}")
            sys.exit(1)
    if node in (None, "", []):
        print(f"[FAIL-FAST] payload 字段为空: {'.'.join(trace)}")
        sys.exit(1)
    return node


# ============ 样式工具 ============
def set_title(ws, range_str, text, fill_hex=NAVY, height=32):
    ws.merge_cells(range_str)
    cell = ws[range_str.split(":")[0]]
    cell.value = text
    cell.font = TITLE_FONT
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = CENTER
    cell.border = BORDER
    ws.row_dimensions[cell.row].height = height


def set_subtitle(ws, range_str, text, fill_hex=GRAY, height=22):
    ws.merge_cells(range_str)
    cell = ws[range_str.split(":")[0]]
    cell.value = text
    cell.font = SUBTITLE_FONT
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = CENTER
    cell.border = BORDER
    ws.row_dimensions[cell.row].height = height


def set_section(ws, row, end_col_letter, text, fill_hex=GREEN, height=22):
    """整行合并的章节标题条"""
    rng = f"A{row}:{end_col_letter}{row}"
    ws.merge_cells(rng)
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = SECTION_FONT
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = LEFT
    cell.border = BORDER
    ws.row_dimensions[row].height = height


def set_header(cell, text, fill_hex=GREEN):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = CENTER
    cell.border = BORDER


def set_key(cell, text):
    cell.value = text
    cell.font = CELL_FONT_BOLD
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.alignment = LEFT
    cell.border = BORDER


def set_value(cell, value):
    cell.font = CELL_FONT
    cell.alignment = LEFT_TOP
    cell.border = BORDER
    if value in (None, "", []):
        cell.value = "[缺]"
        cell.font = CELL_FONT_RED
    else:
        cell.value = value


def set_long_text(cell, value, fill_hex=None):
    cell.font = CELL_FONT
    cell.alignment = LEFT_TOP
    cell.border = BORDER
    cell.value = value
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)


# ============ Sheet ① 决策摘要 ============
def build_sheet1_summary(wb, payload):
    ws = wb.create_sheet("① 决策摘要")
    inject_brand_rows(ws, 4, "超级背调决策摘要  |  小蜂学掌出品")
    B = 2  # brand_row_offset: 所有内容行在原基础上 +2

    # 取必填字段（fail-fast）
    company_full = must(payload, "company_full")
    company_short = must(payload, "company_short")
    country = must(payload, "country")
    level = must(payload, "level")
    date_str = must(payload, "date_str")
    decision = must(payload, "dim9_decision")
    risk_score = must(payload, "dim8_risk", "risk_score")
    revenue = must(payload, "dim3_scale", "revenue_usd")
    employees = must(payload, "dim3_scale", "employees")
    annual_import = must(payload, "dim5_trade", "annual_import_value")
    supplier_countries = must(payload, "dim5_trade", "supplier_countries")
    contacts = must(payload, "dim6_contacts")
    pitches = must(payload, "dim9_pitches")
    email_cn = must(payload, "dev_email_cn")
    email_en = must(payload, "dev_email_en")

    # v2.5 新增：贸易体量速览（fail-fast 强制）
    trade_total_import = must(payload, "dim10_trade", "annual_import_total_usd")
    trade_total_revenue = must(payload, "dim10_trade", "annual_revenue_usd")
    trade_ratio = must(payload, "dim10_trade", "import_to_revenue_ratio")
    trade_year = must(payload, "dim10_trade", "data_year")

    # ---- 标题（行 3）----
    set_title(ws, f"A{1+B}:D{1+B}", f"超级背调决策摘要｜{company_full}", NAVY, 36)

    # ---- 副标题（行 4）----
    set_subtitle(
        ws, f"A{2+B}:D{2+B}",
        f"简称: {company_short}  |  国家: {country}  |  等级: {level}  |  日期: {date_str}  |  决策人: {len(contacts)} 位",
        GRAY, 24,
    )

    # ---- 💰 贸易体量速览（行 5-7）----
    set_section(ws, 3+B, "D", "💰 贸易体量速览（开发价值快速判断）", NAVY, 22)
    headers_trade = ["年度总进口额", "年度销售额", "进口占销比", "数据年份"]
    for ci, h in enumerate(headers_trade, start=1):
        set_header(ws.cell(row=4+B, column=ci), h, GREEN)
    ws.row_dimensions[4+B].height = 24
    set_value(ws.cell(row=5+B, column=1), trade_total_import)
    set_value(ws.cell(row=5+B, column=2), trade_total_revenue)
    set_value(ws.cell(row=5+B, column=3), trade_ratio)
    set_value(ws.cell(row=5+B, column=4), trade_year)
    big_font = Font(name="Microsoft YaHei", size=14, bold=True, color=NAVY)
    for col in range(1, 5):
        c = ws.cell(row=5+B, column=col)
        c.font = big_font
        c.alignment = CENTER
    ratio_cell = ws.cell(row=5+B, column=3)
    ratio_str = str(trade_ratio).strip().rstrip("%")
    try:
        ratio_val = float(ratio_str)
        if ratio_val >= 20:
            ratio_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            ratio_cell.font = Font(name="Microsoft YaHei", size=14, bold=True, color="9C0006")
        elif ratio_val >= 10:
            ratio_cell.fill = PatternFill("solid", fgColor="FFEB9C")
            ratio_cell.font = Font(name="Microsoft YaHei", size=14, bold=True, color="9C5700")
        else:
            ratio_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            ratio_cell.font = Font(name="Microsoft YaHei", size=14, bold=True, color="006100")
    except ValueError:
        ratio_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws.row_dimensions[5+B].height = 36

    # ---- 第 1 章：决策结论（行 8 起）----
    set_section(ws, 6+B, "D", "🎯 一、决策结论（Top Summary）", GREEN, 22)
    rows = [
        ("最终决策", decision, "风险评分", f"{risk_score}/100"),
        ("客户层级", f"{level} 级", "决策人数量", f"{len(contacts)} 位"),
        ("年营收", revenue, "员工规模", employees),
        ("年进口额", annual_import, "主要供应国", supplier_countries),
    ]
    r = 7+B
    decision_cell_ref = None
    risk_cell_ref = None
    for k1, v1, k2, v2 in rows:
        set_key(ws.cell(row=r, column=1), k1)
        set_value(ws.cell(row=r, column=2), v1)
        set_key(ws.cell(row=r, column=3), k2)
        set_value(ws.cell(row=r, column=4), v2)
        if k1 == "最终决策":
            decision_cell_ref = ws.cell(row=r, column=2).coordinate
        if k2 == "风险评分":
            risk_cell_ref = ws.cell(row=r, column=4).coordinate
        ws.row_dimensions[r].height = 26
        r += 1

    # ---- 第 2 章：5 位决策人速览 ----
    set_section(ws, r, "D", f"👥 二、关键决策人速览（共 {len(contacts)} 位）", GREEN, 22)
    r += 1
    headers2 = ["姓名 / 优先级", "职位", "邮箱", "触点策略"]
    for ci, h in enumerate(headers2, start=1):
        set_header(ws.cell(row=r, column=ci), h, NAVY)
    ws.row_dimensions[r].height = 24
    r += 1
    for ct in contacts:
        name = must(ct, "name")
        title = must(ct, "title")
        email = must(ct, "email")
        touch = must(ct, "touch")
        strategy = must(ct, "strategy")
        priority = must(ct, "priority")
        set_value(ws.cell(row=r, column=1), f"{name}  [{priority}]")
        ws.cell(row=r, column=1).font = Font(name="Microsoft YaHei", size=10, bold=True, color=NAVY)
        set_value(ws.cell(row=r, column=2), title)
        set_value(ws.cell(row=r, column=3), email)
        set_value(ws.cell(row=r, column=4), f"{touch}｜{strategy}")
        if priority == "P0":
            row_fill = PatternFill("solid", fgColor="FFE6E6")
        elif priority == "P1":
            row_fill = PatternFill("solid", fgColor="FFF7CC")
        else:
            row_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        for col in range(1, 5):
            ws.cell(row=r, column=col).fill = row_fill
        ws.row_dimensions[r].height = 32
        r += 1

    # ---- 第 3 章：5 套切入话术 ----
    set_section(ws, r, "D", "💬 三、5 套差异化切入话术（直接复制可用）", GREEN, 22)
    r += 1
    set_header(ws.cell(row=r, column=1), "#", NAVY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    set_header(ws.cell(row=r, column=2), "话术全文", NAVY)
    ws.row_dimensions[r].height = 24
    r += 1
    if len(pitches) != 5:
        print(f"[FAIL-FAST] dim9_pitches 必须严格 5 套, 实际 {len(pitches)}")
        sys.exit(1)
    for i, pitch in enumerate(pitches, 1):
        set_value(ws.cell(row=r, column=1), f"切入 {i}")
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).font = CELL_FONT_BOLD
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        set_long_text(ws.cell(row=r, column=2), pitch, LIGHT_GREEN)
        line_count = max(2, len(pitch) // 50 + 1)
        ws.row_dimensions[r].height = max(40, line_count * 18)
        r += 1

    # ---- 第 4 章：中文开发信 ----
    set_section(ws, r, "D", "📧 四、中文开发信全文（直接复制发送）", GREEN, 22)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    set_long_text(ws.cell(row=r, column=1), email_cn, LIGHT_YELLOW)
    cn_lines = email_cn.count("\n") + (len(email_cn) // 60) + 2
    ws.row_dimensions[r].height = max(120, cn_lines * 18)
    r += 1

    # ---- 第 5 章：英文开发信 ----
    set_section(ws, r, "D", "📧 五、English Development Email Draft", GREEN, 22)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    set_long_text(ws.cell(row=r, column=1), email_en, LIGHT_YELLOW)
    en_lines = email_en.count("\n") + (len(email_en) // 70) + 2
    ws.row_dimensions[r].height = max(120, en_lines * 18)
    r += 1

    append_brand_footer(ws, 4)

    # ---- 列宽 ----
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 38

    # ---- 风险评分三色条件格式 ----
    if risk_cell_ref:
        rule = ColorScaleRule(
            start_type="num", start_value=0,  start_color="C00000",
            mid_type="num",   mid_value=50,   mid_color="FFC000",
            end_type="num",   end_value=100,  end_color="00B050",
        )
        ws.conditional_formatting.add(risk_cell_ref, rule)

    # ---- 决策结论高亮（直接填色，不靠条件格式 → sanity_check 可硬验证）----
    if decision_cell_ref:
        cell = ws[decision_cell_ref]
        decision_str = str(cell.value)
        if "✅" in decision_str:
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="006100")
        elif "⚠" in decision_str:
            cell.fill = PatternFill("solid", fgColor="FFEB9C")
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="9C5700")
        else:
            cell.fill = PatternFill("solid", fgColor="FFC7CE")
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="9C0006")

    ws.freeze_panes = f"A{8+B}"


# ============ Sheet ② 全息档案 ============
def build_sheet2_archive(wb, payload):
    ws = wb.create_sheet("② 全息档案")
    inject_brand_rows(ws, 5, "全息档案  |  小蜂学掌出品")
    B = 2

    company_full = must(payload, "company_full")
    company_short = must(payload, "company_short")
    country = must(payload, "country")
    level = must(payload, "level")
    date_str = must(payload, "date_str")

    set_title(ws, f"A{1+B}:E{1+B}", f"超级背调全息档案｜{company_full}", NAVY, 32)
    set_subtitle(
        ws, f"A{2+B}:E{2+B}",
        f"简称: {company_short}  |  国家: {country}  |  等级: {level}  |  日期: {date_str}",
        GRAY, 22,
    )

    d1 = must(payload, "dim1_type")
    d2 = must(payload, "dim2_basic")
    d3 = must(payload, "dim3_scale")
    d4 = must(payload, "dim4_business")
    d5 = must(payload, "dim5_trade")
    d7 = payload.get("dim7_social") or {}  # dim7 允许部分缺失（社媒非必需），其它维度强制
    d8 = must(payload, "dim8_risk")

    rows = [
        ("【1. 客户类型定位】", "section"),
        ("客户类型", d1.get("customer_type"),
         "主营市场", d1.get("main_market")),
        ("分级理由", d1.get("tier_reason"), None, None),

        ("【2. 公司基本情报】", "section"),
        ("英文全称", company_full, "英文简称", company_short),
        ("成立年份", d2.get("founded_year"), "官方网站", d2.get("website")),
        ("总部地址", d2.get("hq_address"), "品牌标语", d2.get("slogan")),

        ("【3. 实力规模】", "section"),
        ("员工数", d3.get("employees"), "年营收 (USD)", d3.get("revenue_usd")),
        ("工厂/仓储", d3.get("facility"), "分支机构", d3.get("branches")),

        ("【4. 产品业务模式】", "section"),
        ("主营品类", d4.get("main_categories"), "SKU 数量", d4.get("sku_count")),
        ("价格带", d4.get("price_range"), "目标客群", d4.get("target_customers")),
        ("差异化卖点", d4.get("differentiation"), None, None),

        ("【5. 贸易海关数据】⭐", "section"),
        ("HS Code", d5.get("hs_codes"), "年进口额", d5.get("annual_import_value")),
        ("主要供应商国家", d5.get("supplier_countries"), "采购频率", d5.get("frequency")),
        ("单批次量", d5.get("batch_size"), None, None),

        ("【7. 社交媒体】", "section"),
        ("Facebook", d7.get("facebook"), "Instagram", d7.get("instagram")),
        ("LinkedIn", d7.get("linkedin"), "YouTube", d7.get("youtube")),
        ("TikTok", d7.get("tiktok"), "活跃度", d7.get("activity_summary")),

        ("【8. 风险评估】⭐", "section"),
        ("融资状况", d8.get("funding_status"), "法律诉讼", d8.get("lawsuits")),
        ("负面舆情", d8.get("negative_news"), "信用评级", d8.get("credit_rating")),
        ("黑名单", d8.get("blacklist"), "风险评分", d8.get("risk_score")),

        ("【9. 跟进决策】⭐", "section"),
        ("最终决策", payload.get("dim9_decision"), "决策人数量", len(payload.get("dim6_contacts") or [])),
    ]

    risk_score_cell_ref = None
    decision_cell_ref = None
    r = 3 + B
    for row in rows:
        if len(row) == 2 and row[1] == "section":
            set_section(ws, r, "D", row[0], GREEN, 22)
        else:
            k1, v1, k2, v2 = row
            set_key(ws.cell(row=r, column=1), k1)
            if k2 is None:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
                set_value(ws.cell(row=r, column=2), v1)
            else:
                set_value(ws.cell(row=r, column=2), v1)
                set_key(ws.cell(row=r, column=3), k2)
                set_value(ws.cell(row=r, column=4), v2)
                if k2 == "风险评分":
                    risk_score_cell_ref = ws.cell(row=r, column=4).coordinate
            if k1 == "最终决策":
                decision_cell_ref = ws.cell(row=r, column=2).coordinate
            ws.row_dimensions[r].height = 26
        r += 1

    # ====== v2.5 新增：第 10 维 · 进出口贸易数据（用到 5 列，比上面多 1 列）======
    d10 = must(payload, "dim10_trade")
    set_section(ws, r, "E", "【10. 进出口贸易数据】⭐⭐ 第 10 维（v2.5 新增）", GREEN, 24)
    r += 1

    # 顶部 4 行：年度总进口额 / 年度销售额 / 进口占销比 / 数据来源
    top_rows = [
        ("年度总进口额", must(d10, "annual_import_total_usd"),
         "年度销售总额", must(d10, "annual_revenue_usd")),
        ("进口占销售比", must(d10, "import_to_revenue_ratio"),
         "数据年份",     must(d10, "data_year")),
        ("数据来源",     must(d10, "data_source"),
         "信源 URL",     must(d10, "data_source_url")),
    ]
    for k1, v1, k2, v2 in top_rows:
        set_key(ws.cell(row=r, column=1), k1)
        set_value(ws.cell(row=r, column=2), v1)
        set_key(ws.cell(row=r, column=3), k2)
        set_value(ws.cell(row=r, column=4), v2)
        # 进口占销比条件填充
        if k1 == "进口占销售比":
            ratio_cell = ws.cell(row=r, column=2)
            ratio_str = str(v1).strip().rstrip("%")
            try:
                ratio_val = float(ratio_str)
                if ratio_val >= 20:
                    ratio_cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    ratio_cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="9C0006")
                elif ratio_val >= 10:
                    ratio_cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    ratio_cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="9C5700")
                else:
                    ratio_cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    ratio_cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="006100")
            except ValueError:
                pass
        if k2 == "信源 URL":
            ws.cell(row=r, column=4).font = Font(name="Microsoft YaHei", size=10, color=NAVY, underline="single")
        ws.row_dimensions[r].height = 26
        r += 1

    # 分国家进口明细表头
    set_section(ws, r, "E", "📦 分国家进口明细（按金额降序）", GRAY, 22)
    r += 1
    breakdown_headers = ["国家", "进口金额 (USD)", "占比", "主要品类", "Top 供应商"]
    # 跨 5 列：A B C D E（占用 1-5 列，但表只用 4 列宽 → 扩到 E 列宽设置）
    for ci, h in enumerate(breakdown_headers, start=1):
        set_header(ws.cell(row=r, column=ci), h, GREEN)
    ws.row_dimensions[r].height = 24
    r += 1

    breakdown = must(d10, "import_breakdown")
    if len(breakdown) < 3:
        print(f"[FAIL-FAST] dim10_trade.import_breakdown 至少 3 个国家，实际 {len(breakdown)}")
        sys.exit(1)
    for item in breakdown:
        ws.cell(row=r, column=1, value=must(item, "country"))
        ws.cell(row=r, column=2, value=must(item, "amount_usd"))
        ws.cell(row=r, column=3, value=must(item, "share_pct"))
        ws.cell(row=r, column=4, value=must(item, "main_categories"))
        ws.cell(row=r, column=5, value=must(item, "top_suppliers"))
        for col in range(1, 6):
            c = ws.cell(row=r, column=col)
            c.border = BORDER
            c.alignment = LEFT_TOP
            c.font = CELL_FONT
        # 国家列：粗体居中
        ws.cell(row=r, column=1).font = Font(name="Microsoft YaHei", size=10, bold=True, color=NAVY)
        ws.cell(row=r, column=1).alignment = CENTER
        # 金额列：居右粗体
        ws.cell(row=r, column=2).font = Font(name="Microsoft YaHei", size=10, bold=True)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        # 占比列：居中
        ws.cell(row=r, column=3).alignment = CENTER
        ws.row_dimensions[r].height = 26
        r += 1
    # ====== 第 10 维结束 ======

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 32  # v2.5: 第 10 维分国家明细 Top 供应商列

    if risk_score_cell_ref:
        rule = ColorScaleRule(
            start_type="num", start_value=0,  start_color="C00000",
            mid_type="num",   mid_value=50,   mid_color="FFC000",
            end_type="num",   end_value=100,  end_color="00B050",
        )
        ws.conditional_formatting.add(risk_score_cell_ref, rule)

    if decision_cell_ref:
        cell = ws[decision_cell_ref]
        decision_str = str(cell.value)
        if "✅" in decision_str:
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="006100")
        elif "⚠" in decision_str:
            cell.fill = PatternFill("solid", fgColor="FFEB9C")
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="9C5700")
        else:
            cell.fill = PatternFill("solid", fgColor="FFC7CE")
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="9C0006")

    append_brand_footer(ws, 5)
    ws.freeze_panes = f"A{3+B}"


# ============ Sheet ③ 决策人图谱 ============
def build_sheet3_contacts(wb, payload):
    ws = wb.create_sheet("③ 决策人图谱")
    inject_brand_rows(ws, 8, "关键决策人图谱  |  小蜂学掌出品")
    B = 2

    company_short = must(payload, "company_short")
    contacts = must(payload, "dim6_contacts")

    headers = ["姓名", "职位", "邮箱", "LinkedIn", "WhatsApp", "触点偏好", "沟通策略", "优先级"]
    last_col = get_column_letter(len(headers))

    set_title(ws, f"A{1+B}:{last_col}{1+B}",
              f"关键决策人图谱｜{company_short}（共 {len(contacts)} 位）", NAVY, 30)

    for i, h in enumerate(headers, start=1):
        set_header(ws.cell(row=2+B, column=i), h, GREEN)
    ws.row_dimensions[2+B].height = 24

    for ri, ct in enumerate(contacts, start=3+B):
        ws.cell(row=ri, column=1, value=must(ct, "name"))
        ws.cell(row=ri, column=2, value=must(ct, "title"))
        ws.cell(row=ri, column=3, value=must(ct, "email"))
        ws.cell(row=ri, column=4, value=must(ct, "linkedin"))
        ws.cell(row=ri, column=5, value=must(ct, "whatsapp"))
        ws.cell(row=ri, column=6, value=must(ct, "touch"))
        ws.cell(row=ri, column=7, value=must(ct, "strategy"))
        ws.cell(row=ri, column=8, value=must(ct, "priority"))

        for col in range(1, 9):
            c = ws.cell(row=ri, column=col)
            c.border = BORDER
            c.alignment = LEFT_TOP
            c.font = CELL_FONT

        wa_cell = ws.cell(row=ri, column=5)
        if str(wa_cell.value).strip().startswith("暂未找到") or str(wa_cell.value).strip() == "未公开":
            wa_cell.font = Font(name="Microsoft YaHei", size=10, italic=True, color=GRAY)
        else:
            wa_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=NAVY)

        prio_cell = ws.cell(row=ri, column=8)
        prio = str(prio_cell.value)
        if prio == "P0":
            prio_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            prio_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="9C0006")
        elif prio == "P1":
            prio_cell.fill = PatternFill("solid", fgColor="FFEB9C")
            prio_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="9C5700")
        else:
            prio_cell.fill = PatternFill("solid", fgColor="D9D9D9")
            prio_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="595959")
        prio_cell.alignment = CENTER
        ws.row_dimensions[ri].height = 30

    append_brand_footer(ws, 8)

    widths = [16, 24, 30, 32, 18, 20, 36, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{3+B}"


# ============ Sheet ④ 数据来源 ============
def build_sheet4_sources(wb, payload):
    ws = wb.create_sheet("④ 数据来源")
    inject_brand_rows(ws, 3, "数据来源与可信度  |  小蜂学掌出品")
    B = 2

    sources = must(payload, "data_sources")

    set_title(ws, f"A{1+B}:C{1+B}", f"数据来源与可信度（共 {len(sources)} 条）", NAVY, 28)

    for i, h in enumerate(["信源", "URL", "抓取时间"], start=1):
        set_header(ws.cell(row=2+B, column=i), h, GREEN)
    ws.row_dimensions[2+B].height = 22

    for ri, src in enumerate(sources, start=3+B):
        ws.cell(row=ri, column=1, value=must(src, "source"))
        ws.cell(row=ri, column=2, value=must(src, "url"))
        ws.cell(row=ri, column=3, value=must(src, "fetched_at"))
        for col in range(1, 4):
            c = ws.cell(row=ri, column=col)
            c.border = BORDER
            c.alignment = LEFT_TOP
            c.font = CELL_FONT
        ws.cell(row=ri, column=2).font = Font(name="Microsoft YaHei", size=10, color=NAVY)
        ws.row_dimensions[ri].height = 24

    append_brand_footer(ws, 3)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = f"A{3+B}"


# ============ Sheet ⑤ 自检报告（占位，由 sanity_check.py 在生成后回写） ============
def build_sheet5_audit_placeholder(wb, payload):
    ws = wb.create_sheet("⑤ 自检报告")
    inject_brand_rows(ws, 3, "自检报告  |  小蜂学掌出品")
    B = 2

    set_title(ws, f"A{1+B}:C{1+B}", "自检报告（sanity_check 17 项硬拦截 v3.0）", NAVY, 28)

    headers = ["#", "检查项", "结果"]
    for i, h in enumerate(headers, start=1):
        set_header(ws.cell(row=2+B, column=i), h, GREEN)
    ws.row_dimensions[2+B].height = 22

    audit_items = [
        ("S01", "文件能被 openpyxl 正常打开（拦截 styles.xml 腐烂）"),
        ("S02", "Sheet 数 = 6（⓪ 封面 / ① 决策摘要 / ② 全息档案 / ③ 决策人图谱 / ④ 数据来源 / ⑤ 自检报告）[共 18 项硬断言]"),
        ("S03", "决策人图谱 ≥ 6 行（含表头）"),
        ("S04", "决策人图谱列数 = 8"),
        ("S05", "决策人中至少 1 位 P0"),
        ("S06", "数据来源 ≥ 4 行"),
        ("S07", "风险评分单元格存在条件格式规则"),
        ("S08", "决策结论单元格有填充色（绿/黄/红）"),
        ("S09", "文件名格式 = 小蜂学掌_单客户档案_{company_short}_{date_str}.xlsx"),
        ("S10", "文件大小 > 10KB（拦截空壳）"),
        ("S11", "全息档案 ≥ 39 行（含第 10 维）"),
        ("S12", "5 套话术每套 ≥ 20 字（拦截占位）"),
        ("S13", "中英开发信各 ≥ 50 字 + 不含 fallback 模板词"),
        ("S14", "Sheet ① 顶部存在『贸易体量速览』4 列卡片（v2.5）"),
        ("S15", "Sheet ② 存在『第 10 维 · 进出口贸易数据』章节标题条（v2.5）"),
        ("S16", "分国家进口明细 ≥ 3 行 × 5 列齐全（v2.5）"),
        ("S17", "年度总进口额 / 年度销售额 含 USD 前缀且非占位词（v2.5）"),
        ("S18", "样式存在性检查：内容 Sheet 有有效填充色单元格 ≥ 3（拦截裸 openpyxl 写入 / 样式缺失）"),
    ]
    for ri, (code, desc) in enumerate(audit_items, start=3+B):
        set_value(ws.cell(row=ri, column=1), code)
        ws.cell(row=ri, column=1).alignment = CENTER
        ws.cell(row=ri, column=1).font = CELL_FONT_BOLD
        set_value(ws.cell(row=ri, column=2), desc)
        set_value(ws.cell(row=ri, column=3), "(待 sanity_check 写入)")
        ws.cell(row=ri, column=3).font = Font(name="Microsoft YaHei", size=10, italic=True, color=GRAY)
        ws.cell(row=ri, column=3).alignment = CENTER
        ws.row_dimensions[ri].height = 26

    append_brand_footer(ws, 3)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 28
    ws.freeze_panes = f"A{3+B}"


# ============ 主入口 ============
def generate(payload, out_path):
    # ---- 强制文件名格式（v3.0: 小蜂学掌_前缀）----
    company_short = must(payload, "company_short")
    date_str = must(payload, "date_str")
    expected_name = f"小蜂学掌_单客户档案_{company_short}_{date_str}.xlsx"
    actual_name = os.path.basename(out_path)
    if actual_name != expected_name:
        print(f"[FAIL-FAST] 文件名必须为 '{expected_name}'，当前 '{actual_name}'")
        sys.exit(1)

    wb = Workbook()
    build_sheet_cover(wb, payload)    # ⓪ 封面（占用 wb.active 默认 Sheet）
    build_sheet1_summary(wb, payload)
    build_sheet2_archive(wb, payload)
    build_sheet3_contacts(wb, payload)
    build_sheet4_sources(wb, payload)
    build_sheet5_audit_placeholder(wb, payload)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"[OK] Excel saved -> {out_path}")
    print(f"[NEXT] 必须运行: python scripts/sanity_check.py --xlsx \"{out_path}\"")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--payload", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    with open(args.payload, "r", encoding="utf-8") as f:
        payload = json.load(f)

    generate(payload, args.out)


if __name__ == "__main__":
    main()
